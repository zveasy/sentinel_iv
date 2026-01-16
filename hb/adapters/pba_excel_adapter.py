import csv
import os

from hb.registry_utils import normalize_alias


def _load_rows_csv(path):
    with open(path, "r", newline="") as f:
        reader = csv.reader(f)
        rows = [row for row in reader]
    return rows


def _load_rows_csv_streaming(path):
    with open(path, "r", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            yield row


def _load_rows_xlsx(path):
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    sheets = workbook.worksheets
    unit_map = {}
    for sheet in sheets:
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        if _detect_unit_sheet(rows):
            unit_map = _parse_unit_sheet(rows)
            continue
        header_idx = _detect_table_header(rows)
        if header_idx is not None:
            return rows, unit_map
    return [[cell for cell in row] for row in workbook.active.iter_rows(values_only=True)], unit_map


def _load_rows_xlsx_streaming(path):
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    unit_map = {}
    metrics = {}
    found_data = False

    for sheet in workbook.worksheets:
        header = None
        header_row = None
        header_next = None
        unit_sheet = False

        for row in sheet.iter_rows(values_only=True):
            if row is None or not any(cell is not None for cell in row):
                continue
            if header_row is None:
                header_row = list(row)
                header_norm = [normalize_alias(str(cell)) for cell in header_row if cell is not None]
                if "metric" in header_norm and "unit" in header_norm and "current" not in header_norm and "value" not in header_norm:
                    unit_sheet = True
                if unit_sheet:
                    header = header_row
                else:
                    header = header_row
                continue

            if unit_sheet:
                col_map = {normalize_alias(str(name)): idx for idx, name in enumerate(header) if name}
                metric_idx = col_map.get("metric")
                unit_idx = col_map.get("unit")
                if metric_idx is None or unit_idx is None:
                    continue
                if metric_idx < len(row) and unit_idx < len(row):
                    metric_name = str(row[metric_idx]).strip()
                    unit_name = str(row[unit_idx]).strip()
                    if metric_name:
                        unit_map[metric_name] = unit_name
                continue

            if header_next is None:
                header_next = list(row)
                header_norm = {normalize_alias(str(c)) for c in header if c}
                if "current" not in header_norm and "value" not in header_norm:
                    header = _merge_header_rows(header, header_next)
                found_data = True
                continue

            if not found_data:
                continue

            col_map = {normalize_alias(str(name)): idx for idx, name in enumerate(header) if name}
            if "metric" not in col_map:
                raise ValueError("missing required column: Metric")
            metric_idx = col_map.get("metric")
            if metric_idx is None or metric_idx >= len(row):
                continue
            metric_name = str(row[metric_idx]).strip()
            if metric_name == "":
                continue
            current_idx = col_map.get("current")
            value_idx = col_map.get("value")
            unit_idx = col_map.get("unit")
            if current_idx is None and value_idx is None:
                raise ValueError("missing required column: Current or Value")
            value = None
            if current_idx is not None and current_idx < len(row):
                value = row[current_idx]
            elif value_idx is not None and value_idx < len(row):
                value = row[value_idx]
            unit = None
            if unit_idx is not None and unit_idx < len(row):
                unit = row[unit_idx]
            metrics[metric_name] = {"value": value, "unit": unit}

    # Fill units from unit_map if missing.
    for metric_name, data in metrics.items():
        if data.get("unit") is None and metric_name in unit_map:
            data["unit"] = unit_map[metric_name]

    return metrics


def _detect_table_header(rows):
    for idx, row in enumerate(rows):
        if not row:
            continue
        if any(str(cell).strip().lower() == "metric" for cell in row if cell is not None):
            return idx
    return None


def _detect_unit_sheet(rows):
    if not rows:
        return False
    header = rows[0]
    header_norm = [normalize_alias(str(cell)) for cell in header if cell is not None]
    if "metric" in header_norm and "unit" in header_norm:
        if "current" not in header_norm and "value" not in header_norm:
            return True
    return False


def _parse_unit_sheet(rows):
    if not rows:
        return {}
    header = rows[0]
    col_map = {normalize_alias(str(name)): idx for idx, name in enumerate(header) if name}
    metric_idx = col_map.get("metric")
    unit_idx = col_map.get("unit")
    if metric_idx is None or unit_idx is None:
        return {}
    unit_map = {}
    for row in rows[1:]:
        if not row:
            continue
        if metric_idx >= len(row) or unit_idx >= len(row):
            continue
        metric_name = str(row[metric_idx]).strip()
        unit_name = str(row[unit_idx]).strip()
        if metric_name:
            unit_map[metric_name] = unit_name
    return unit_map


def _merge_header_rows(primary, secondary):
    merged = []
    for idx, value in enumerate(primary):
        if value is None or str(value).strip() == "":
            merged.append(secondary[idx] if idx < len(secondary) else value)
        else:
            merged.append(value)
    return merged


def parse(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".csv"]:
        rows = _load_rows_csv(path)
        unit_map = {}
    else:
        rows, unit_map = _load_rows_xlsx(path)

    if not rows:
        raise ValueError("source contains no rows")

    header_idx = _detect_table_header(rows)
    if header_idx is not None:
        header = rows[header_idx]
        if header_idx + 1 < len(rows):
            next_row = rows[header_idx + 1]
            if header and next_row and (
                "current" not in {normalize_alias(str(c)) for c in header if c}
                and "value" not in {normalize_alias(str(c)) for c in header if c}
            ):
                header = _merge_header_rows(header, next_row)
                data_rows = rows[header_idx + 2 :]
            else:
                data_rows = rows[header_idx + 1 :]
        else:
            data_rows = rows[header_idx + 1 :]
        col_map = {normalize_alias(str(name)): idx for idx, name in enumerate(header) if name}
        metrics = {}
        if "metric" not in col_map:
            raise ValueError("missing required column: Metric")
        for row in data_rows:
            if not row:
                continue
            metric_idx = col_map.get("metric")
            if metric_idx is None or metric_idx >= len(row):
                continue
            metric_name = str(row[metric_idx]).strip()
            if metric_name == "":
                continue
            current_idx = col_map.get("current")
            value_idx = col_map.get("value")
            unit_idx = col_map.get("unit")
            if current_idx is None and value_idx is None:
                raise ValueError("missing required column: Current or Value")
            value = None
            if current_idx is not None and current_idx < len(row):
                value = row[current_idx]
            elif value_idx is not None and value_idx < len(row):
                value = row[value_idx]
            unit = None
            if unit_idx is not None and unit_idx < len(row):
                unit = row[unit_idx]
            if unit is None and metric_name in unit_map:
                unit = unit_map[metric_name]
            metrics[metric_name] = {"value": value, "unit": unit}
        return metrics

    # Fallback: treat as metric,value table
    metrics = {}
    for row in rows:
        if not row or len(row) < 2:
            continue
        metric_name = str(row[0]).strip()
        if metric_name == "":
            continue
        unit = None
        if metric_name in unit_map:
            unit = unit_map[metric_name]
        metrics[metric_name] = {"value": row[1], "unit": unit}
    return metrics


def parse_stream(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xlsm"]:
        return _load_rows_xlsx_streaming(path)
    elif ext in [".csv"]:
        rows = None
        unit_map = {}
    else:
        return parse(path)

    metrics = {}
    header = None
    data_rows = []
    if rows is None:
        for row in _load_rows_csv_streaming(path):
            if not row:
                continue
            if header is None and any(str(cell).strip().lower() == "metric" for cell in row if cell is not None):
                header = row
                continue
            if header is not None:
                data_rows.append(row)

    if header is None:
        # Fallback: metric,value rows
        for row in _load_rows_csv_streaming(path):
            if not row or len(row) < 2:
                continue
            metric_name = str(row[0]).strip()
            if metric_name == "":
                continue
            metrics[metric_name] = {"value": row[1], "unit": None}
        return metrics

    col_map = {normalize_alias(str(name)): idx for idx, name in enumerate(header) if name}
    if "metric" not in col_map:
        raise ValueError("missing required column: Metric")
    for row in data_rows:
        metric_idx = col_map.get("metric")
        if metric_idx is None or metric_idx >= len(row):
            continue
        metric_name = str(row[metric_idx]).strip()
        if metric_name == "":
            continue
        current_idx = col_map.get("current")
        value_idx = col_map.get("value")
        unit_idx = col_map.get("unit")
        if current_idx is None and value_idx is None:
            raise ValueError("missing required column: Current or Value")
        value = None
        if current_idx is not None and current_idx < len(row):
            value = row[current_idx]
        elif value_idx is not None and value_idx < len(row):
            value = row[value_idx]
        unit = None
        if unit_idx is not None and unit_idx < len(row):
            unit = row[unit_idx]
        if unit is None and metric_name in unit_map:
            unit = unit_map[metric_name]
        metrics[metric_name] = {"value": value, "unit": unit}
    return metrics
